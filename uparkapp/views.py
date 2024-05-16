from django.utils import timezone
from .functions import *
from typing import Any
from django.utils.datastructures import MultiValueDictKeyError
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.core.exceptions import *
from django.http import HttpRequest, HttpResponse
from django.views.generic import ListView
from django.views.generic.base import TemplateView
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import *
from django.db.models import Max
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, logout, authenticate
import matplotlib.pyplot as plt
import matplotlib
import io
import urllib, base64
import datetime
from datetime import datetime
from flask import Flask, send_file

#Manejo de libros de excel
from openpyxl import Workbook
from openpyxl.styles import Font, Color,colors,Alignment
import xlsxwriter
import pandas as pd



def masive(request):
    mensaje = masivePays()
    return render (request, 'errors.html',{"error": mensaje})

def flatFile(request):
    pays = Pay.objects.all().order_by('date')
    # file = genFlatFile(pays)
    # return HttpResponse('<meta http-equiv="refresh" content="0; /admin"/>')
    file = open("./uPark/media/files/flatFile.txt", "w")
    for p in pays:
        person = Person.objects.get(idPerson = p.idPerson_id)
        file.write(str(p.idPay)+";"+str(p.transactionValue)+";"+str(p.cusCod)+";"+str(p.date.strftime('%Y-%m-%d %H:%M:%S'))+";"+str(person.documentId)+";"+str(p.idVehicle_id)+";"+ "\n")
    file.close()
    
    #file = open("./uPark/media/files/flatFile.txt", "r")
    #content = "attachment; filename={0}".format(file)
    return send_file("./uPark/media/files/flatFile.txt", as_attachment=True)
    #return HttpResponse('<meta http-equiv="refresh" content="0; /admin"/>')

def admin(request):
    return render (request, 'admin.html')

def prueba(request):
    return render (request, 'prueba.html')

def adminuser(request):
    personList = Person.objects.all()
    return render (request, 'adminuser.html', {"Persons": personList})

def card(request):
    cardList= Card.objects.filter(status = 'A').order_by("idCard") 
    #cardList= Card.objects.all().order_by("idCard") 
    return render (request, 'card.html',{"Card": cardList})

def visitor(request):
    return render (request, 'visitor.html')

def vehicle(request):
    vehiclelist=Vehicle.objects.all().order_by("idVehicle")  
    return render(request,'vehicle.html',{"Vehicles":vehiclelist})


def pse(request):    
    return render (request, 'errors.html',{"error":"Vengo de uparkapp"})

def errors(request):
    return render (request, 'errors.html',{"error": request.POST['error']})


def addVehicle (request):
    type = request.POST ['type']
    rate = request.POST ['rate'] 
    vehicle = Vehicle.objects.create(type=type, rate=rate)
    vehicle.save()
    return redirect ('/vehicle')

def editVehicle (request, idVehicle):
    vehicle = Vehicle.objects.get(idVehicle=idVehicle)
    return render (request, "editVehicle.html",{"vehicle" : vehicle})

def editarVehicle (request):
    idVehicle = request.POST ['idVehicle']
    type = request.POST ['type']
    rate = request.POST ['rate'] 
    vehicle = Vehicle.objects.get(idVehicle=idVehicle)
    vehicle.type = type
    vehicle.rate = rate
    vehicle.save()
    return redirect ('/vehicle')    

def deleteVehicle (request, idVehicle):
    vehicle = Vehicle.objects.get(idVehicle=idVehicle)
    vehicle.delete()
    return redirect ('/vehicle')

def addPerson (request):
    mail = str.replace(request.POST ['mail'].lower(), ' ', '')
    user = Person.objects.filter(mail = mail).first()
    if user:
        return render (request, 'errors.html',{"error": "The email is already assigned to another person"})
    else:
        try:
            document = request.POST ['documentId']       
            firstName = request.POST ['firstName']
            lastName = request.POST ['lastName']
            pwd = encryptPwd(request.POST ['password'])
            phone  = request.POST ['phone']
            dateOfBirth = request.POST ['dateOfBirth']
            personType = request.POST ['personType']
            person = Person.objects.create(documentId=document,
                                        firstName=firstName,
                                        lastName=lastName,
                                        password = pwd,
                                        phone=phone,
                                        mail=mail,
                                        dateOfBirth=dateOfBirth,
                                        personType=personType)        
            createCard(person.idPerson, 0)
            person.save()
            return redirect ('/adminuser')
        except:
            return render (request, 'errors.html',{"error": "Something went wrong!"})

def createCard (idPerson, balance):
    card = Card.objects.create (idPerson_id=idPerson, balance = balance,status='A')
    card.save()
    

def editCard (request, idCard):
    card = Card.objects.get(idCard=idCard)
    return render (request, "editCard.html",{"card" : card})

def editarCard (request):
    balance = request.POST ['balance']
    status = request.POST ['status']
    idCard = request.POST ['idCard']
    card = Card.objects.get(idCard=idCard)
    if status == 'A':
        card.balance = balance
        card.save()
    else:
        card.status = status
        card.balance = 0
        card.save()
        createCard(card.idPerson.idPerson, balance)
        
    return redirect ('/card')    
    
def welcome (request):
    try:
        person = Person.objects.get(mail = request.POST['username'])
        #return render(request, "Errors.html",{"error":person.password})
    except:
        return render(request, "errors.html",{"error": "usuario no existe"})    
    if (decryptPwd(person.password, request.POST['password'])):
        card = Card.objects.get(idPerson =  person.idPerson)
        if person.personType == 'A':
            return HttpResponse('<meta http-equiv="refresh" content="0; /admin"/>')
        else:
            paylist=Pay.objects.filter(idPerson_id= person.idPerson).order_by("-date")  
            return render(request, "welcome.html",{"resulPerson": person.firstName, "idPerson":person.idPerson, "resulCard": card.balance, "Viewpay":paylist, "idPersons":person.idPerson})
    else:
        return render(request, "errors.html",{"error": "Usuario o contraseña inválida"})



def reportVehicle(request):
    vehicle = Vehicle.objects.all()
    wb = Workbook()
    ws = wb.active
    font = Font(b=True, color="00000080")
    alignment = Alignment(horizontal="center", vertical="center")
    a1 = ws['A1']
    a1.font = font 
    a1.alignment = alignment
    a1.value  = 'VEHICLE LIST'
    #Cabezera de reporte
    ws.merge_cells('A1:C1')
    a3 = ws['A3']
    b3 = ws['B3'] 
    c3 = ws['C3']
    a3.font = font 
    a3.alignment = alignment
    b3.font = font 
    b3.alignment = alignment
    c3.font = font 
    c3.alignment = alignment    
    a3.value = 'Id Vehicle'
    b3.value = 'Type Vehicle'
    c3.value = 'Rate'
    
    cont = 4
    for vehicle in vehicle:
        ws.cell(row = cont, column =1).value = vehicle.idVehicle
        ws.cell(row = cont, column =2).value = vehicle.type
        ws.cell(row = cont, column =3).value = vehicle.rate
        cont += 1
    fileName= "List_vehicle_uPark.xlsx"
    response = HttpResponse(content_type = "applications/ms-excel")
    content = "attachment; filename={0}".format(fileName)
    response['Content-Disposition'] = content
    wb.save(response)
    return (response)

def reportCard(request):
    card = Card.objects.select_related ('idPerson').all()   
    wb = Workbook()
    ws = wb.active
    font = Font(b=True, color="00000080")
    alignment = Alignment(horizontal="center", vertical="center")
    a1 = ws['A1']
    a1.font = font 
    a1.alignment = alignment
    a1.value  = 'CARD LIST'
    #Cabezera de reporte
    ws.merge_cells('A1:E1')
    a3 = ws['A3']
    b3 = ws['B3']
    c3 = ws['C3']
    d3 = ws['D3']
    e3 = ws['E3']
    a3.font = font 
    a3.alignment = alignment
    b3.font = font 
    b3.alignment = alignment
    c3.font = font 
    c3.alignment = alignment
    d3.font = font 
    d3.alignment = alignment
    e3.font = font 
    e3.alignment = alignment    
    a3.value = 'Id Card'
    b3.value = 'Id Person'
    c3.value = 'Name'
    d3.value = 'Balance'
    e3.value = 'Status'    
    cont = 4
    for card in card:
        ws.cell(row = cont, column =1).value = card.idCard
        ws.cell(row = cont, column =2).value = card.idPerson.idPerson
        ws.cell(row = cont, column =3).value = card.idPerson.firstName+' '+card.idPerson.lastName
        ws.cell(row = cont, column =4).value = card.balance
        if card.status == 'A':
            ws.cell(row = cont, column =5).value = "Active"
        elif card.status == 'I':
            ws.cell(row = cont, column =5).value = "Inactive"
        else:
            ws.cell(row = cont, column =5).value = "Other status"
        #ws.cell(row = cont, column =5).value = card.status
        cont += 1
    fileName= "List_card_uPark.xlsx"

    response = HttpResponse(content_type = "applications/ms-excel")
    content = "attachment; filename={0}".format(fileName)
    response['Content-Disposition'] = content
    wb.save(response)
    return (response)           

def reportPay(request):
    pay = Pay.objects.select_related ('idPerson').all()   
    wb = Workbook()
    ws = wb.active
    font = Font(b=True, color="00000080")
    alignment = Alignment(horizontal="center", vertical="center")
    a1 = ws['A1']
    a1.font = font 
    a1.alignment = alignment
    a1.value  = 'PAY LIST'
    #Cabezera de reporte
    ws.merge_cells('A1:F1')
    a3 = ws['A3']
    b3 = ws['B3']
    c3 = ws['C3']
    d3 = ws['D3']
    e3 = ws['E3']
    f3 = ws['F3']
    a3.font = font 
    a3.alignment = alignment
    b3.font = font 
    b3.alignment = alignment
    c3.font = font 
    c3.alignment = alignment
    d3.font = font 
    d3.alignment = alignment
    e3.font = font 
    e3.alignment = alignment 
    f3.font = font 
    f3.alignment = alignment    
    a3.value = 'Id Pay'
    b3.value = 'Id Person'
    c3.value = 'Name'
    d3.value = 'Vehicle'
    e3.value = 'Value'   
    f3.value = 'Date'    
    cont = 4
    for pay in pay:
        ws.cell(row = cont, column =1).value = pay.idPay
        ws.cell(row = cont, column =2).value = pay.idPerson.idPerson
        ws.cell(row = cont, column =3).value = pay.idPerson.firstName+' '+pay.idPerson.lastName
        ws.cell(row = cont, column =4).value = pay.idVehicle.type
        ws.cell(row = cont, column =5).value = pay.transactionValue
        ws.cell(row = cont, column =6).value = pay.date.strftime("%d/%m/%Y")
        cont += 1
    fileName= "List_pay_uPark.xlsx"

    response = HttpResponse(content_type = "applications/ms-excel")
    content = "attachment; filename={0}".format(fileName)
    response['Content-Disposition'] = content
    wb.save(response)
    return (response)

def reportPerson(request):
    persons = Person.objects.all()   
    wb = Workbook()
    ws = wb.active
    font = Font(b=True, color="00000080")
    alignment = Alignment(horizontal="center", vertical="center")
    a1 = ws['A1']
    a1.font = font 
    a1.alignment = alignment
    a1.value  = 'PERSON LIST'
    #Cabezera de reporte
    ws.merge_cells('A1:E1')
    a3 = ws['A3']
    b3 = ws['B3']
    c3 = ws['C3']
    d3 = ws['D3']
    e3 = ws['E3']
    f3 = ws['F3']
    g3 = ws['G3']
    a3.font = font 
    a3.alignment = alignment
    b3.font = font 
    b3.alignment = alignment
    c3.font = font 
    c3.alignment = alignment
    d3.font = font 
    d3.alignment = alignment
    e3.font = font 
    f3.alignment = alignment
    f3.font = font 
    g3.alignment = alignment   
    g3.font = font 
    e3.alignment = alignment       
    a3.value = 'Id Person'
    b3.value = 'Document ID'
    c3.value = 'Name'
    d3.value = 'Phone'
    e3.value = 'Mail'
    f3.value = 'Date of Birth'
    g3.value = 'Person Type'
    cont = 4
    for per in persons:
        ws.cell(row = cont, column =1).value = per.idPerson
        ws.cell(row = cont, column =2).value = per.documentId
        ws.cell(row = cont, column =3).value = per.firstName+' '+per.lastName
        ws.cell(row = cont, column =4).value = per.phone
        ws.cell(row = cont, column =5).value = per.mail
        ws.cell(row = cont, column =6).value = per.dateOfBirth
        ws.cell(row = cont, column =7).value = per.personType
        cont += 1
    fileName= "List_Persons_uPark.xlsx"

    response = HttpResponse(content_type = "applications/ms-excel")
    content = "attachment; filename={0}".format(fileName)
    response['Content-Disposition'] = content
    wb.save(response)
    return (response)     